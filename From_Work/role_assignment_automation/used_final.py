from openpyxl import Workbook
from openpyxl import load_workbook

def next_col(current):
    if current[-1] == 'Z':
        i = len(current)-1
        while current[i] == 'Z' and i >=0:
            i-=1
        if i == -1:
            return 'A'*(len(current)+1)
        else:
            nxt = current[:i]

            nxt += str(chr(ord(current[i]) + 1))
            to_fill = len(current)-1-i
            for x in range(to_fill):
                nxt += "A"
            return nxt
    else:
        next_val = chr(ord(current[-1]) + 1)
        retval = current[:len(current)-1] + next_val
        return retval


LINES_FOR_CURRENT_EXCEL_FILE = 166

desc_to_row = dict() # get the row to write to in the output file for each permission description

output_file = load_workbook('out_file1.xlsx') # define the output file
ws = output_file.active
line_no = 3
while line_no <= LINES_FOR_CURRENT_EXCEL_FILE:
    if str(ws['A' + str(line_no)].value) != "None": # if there's something in the A column for this row
        desc_to_row[str(ws['A' + str(line_no)].value).rstrip().upper()] = line_no
        print("A", line_no)
    elif str(ws['B' + str(line_no)].value) != "None":
        desc_to_row[str(ws['B' + str(line_no)].value).rstrip().upper()] = line_no
        print("B", line_no)
    elif str(ws['C' + str(line_no)].value) != "None":
        desc_to_row[str(ws['C' + str(line_no)].value).rstrip().upper()] = line_no
        print("C", line_no)

    line_no += 1


#desc_to_row["DRV SPOOLFLEX"] = 233
#desc_to_row["CHANGE CREDIT LIMIT"] = 234
#desc_to_row["IBM QUERY"] = 235
#desc_to_row["DSO ERAM REPORT"] = 237
#desc_to_row["CONTROL FILE MAINTENANCE"] = 3
#desc_to_row["DISPLAY ENTITY CONTROL NUMBERS"] = 33
#desc_to_row["DISPLAY CODE TYPES"] = 34
#desc_to_row["LIST ENTITY CONTROL NUMBERS"] = 58

new_descriptions = [] # [("DRV SPOOLFLEX", 233), ("CHANGE CREDIT LIMIT", 234), ("IBM QUERY", 235), ("DSO ERAM REPORT", 237), ("CONTROL FILE MAINTENANCE",3), ("DISPLAY ENTITY CONTROL NUMBERS", 33), ("DISPLAY CODE TYPES", 34), ("LIST ENTITY CONTROL NUMBERS", 58)]

new_amount_of_lines = 166
def zeros_index(line): # if this line has a description, return the index where '000' occurs
    if "000" in line:
        return line.index("000")
    else:
        return -1

# read the lines from the file
file = open('input_file1.txt', 'r')
lines = [line.strip().split() for line in file.readlines()]
print(lines)

current_line = 0
# get to first 'user' statement
user_name = ""
done = False
while not done and current_line < LINES_FOR_CURRENT_EXCEL_FILE:
    
    if len(lines[current_line]) != 0 and lines[current_line][0] == "User-":
        user_name = lines[current_line][1]
        done = True
    current_line += 1

print(user_name)
print(current_line)

current_col = "H"
while current_line < len(lines):
    # read descriptions until the next user shows up
    next_user = False
    permissions_this_user = [] # the row numbers to write for this user
    while not next_user and current_line < len(lines):
        # check to see if the 'user' identifier is the first word in this line
        if len(lines[current_line]) == 0: # empty line
            current_line += 1
            continue
        if lines[current_line][0] == "User-":
            if lines[current_line][1] != user_name: # it's a new user
                user_just_finished = user_name
                user_name = lines[current_line][1]
                print("New user:", user_name)
                next_user = True
                current_line += 1
                break
            else:
                current_line += 1
                continue
        zeros_idx = zeros_index(lines[current_line])
        if zeros_idx != -1 and zeros_idx != len(lines[current_line])-1: # this line has a permissions description
            description = ""
            for word in lines[current_line][zeros_idx+1:]:
                description += word + " "
            description = description[:len(description)-1] # get rid of the last space
            print("Description:", description)
            to_add = 0
            try:
                to_add = desc_to_row[description.upper()]
            except KeyError:
                to_add = new_amount_of_lines + 1
                new_amount_of_lines += 1
                desc_to_row[description.upper()] = to_add
                new_descriptions.append((description.upper(), to_add))

            permissions_this_user.append(to_add)
            current_line += 1
        else:
            current_line += 1

    if current_line == len(lines):
        user_just_finished = user_name
    # write permissions for this user
    print("Permissions for", user_just_finished, ": ", permissions_this_user)

    print("Current column: ",current_col)
    ws[current_col + '2'] = user_just_finished
    for row_number in permissions_this_user:
        ws[current_col + str(row_number)] = 'x'

    current_col = next_col(current_col)
    print(current_col)


for desc, row_number in new_descriptions:
    ws['A' + str(row_number)] = desc

output_file.save('GL_sheet.xlsx')