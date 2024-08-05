from openpyxl import Workbook
from openpyxl import load_workbook
import fractions
import math

def is_mixed(number):
    return '-' in number

def mixed_to_float(s):
    one, two = s.split('-')
    num, denom = two.split('/')
    num = int(num)
    denom = int(denom)
    num = num + denom*int(one)
    return num / denom

def is_int(number):
    return '/' not in number and '.' not in number

def is_float(number):
    return '/' not in number

data_file = open('Round_Tube.txt', 'r')

wb = load_workbook("test_0.xlsx")
ws = wb["RT=Round Tube"]
#["FB=Flat Bar"]

digits = "0123456789.-—"
current_w1 = 0
current_w1_str = ""
desc_set = set()
current_line = 4
for line in data_file.readlines():
    numbers = line.strip().split()
    if len(numbers) == 0:
        continue
    elif numbers[0][0] not in digits:
        continue
    elif len(numbers) == 1: # new w1
        if is_mixed(numbers[0]):
            current_w1 = mixed_to_float(numbers[0])
            one, two = numbers[0].split('-')
            current_w1_str = one + " " + two
        elif is_int(numbers[0]) or is_float(numbers[0]):
            current_w1 = float(numbers[0])
            current_w1_str = numbers[0]
        else:
            num, denom = numbers[0].split('/')
            current_w1 = int(num)/int(denom)
            current_w1_str = str(current_w1)
        continue

    # read the line
    if is_mixed(numbers[1]):
        numbers[1] = str(mixed_to_float(numbers[1]))
    elif '/' in numbers[1]:
        num, denom = numbers[1].split('/')
        numbers[1] = str(int(num)/int(denom))

    print(numbers)
    desc = "Round Tube " + str(current_w1) + " x " + numbers[2] + " x " + numbers[1]

    if desc not in desc_set:
        desc_set.add(desc)
        ws["A" + str(current_line)] = desc
        ws["D" + str(current_line)] = numbers[1]
        ws["F" + str(current_line)] = current_w1
        ws["G" + str(current_line)] = numbers[2]
        ws["H" + str(current_line)] = numbers[3]
        current_line += 1

wb.save("Round Tube0 Complete.xlsx")

# round_tube0 is for K4-64
# w1
# junk th1 w2 wgt junk junk

# round tube1 is for K118-123. 
# w1
#  th1 w2 wgt junk junk

# round tube2 for K128
# junk w1 th1 w2 wgt

# round tube3 for K133
# w1 th1 w2 wgt