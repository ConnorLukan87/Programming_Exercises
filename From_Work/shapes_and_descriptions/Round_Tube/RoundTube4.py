
from openpyxl import Workbook
from openpyxl import load_workbook
import fractions
import math

def is_mixed(number):
    return '-' in number

def mixed_to_float(s):
    one, two = s.split('-')
    num, denom = two.split('/')
    num = int(num)
    denom = int(denom)
    num = num + denom*int(one)
    return num / denom

def is_int(number):
    return '/' not in number and '.' not in number

def is_float(number):
    return '/' not in number

data_file = open('Round_Tube5.txt', 'r')

wb = load_workbook("test_0.xlsx")
ws = wb["RT=Round Tube"]
#["FB=Flat Bar"]

digits = "0123456789.-—"
desc_set = set()
current_line = 4
for line in data_file.readlines():
    numbers = line.strip().split()
    desc = "Round Tube " + numbers[0] + " x " + numbers[1] + " x " + numbers[2]
    if desc not in desc_set:
        desc_set.add(desc)
        ws["A" + str(current_line)] = desc
        ws["D" + str(current_line)] = numbers[2]
        ws["F" + str(current_line)] = numbers[0]
        ws["G" + str(current_line)] = numbers[1]
        ws["H" + str(current_line)] = numbers[3]
        current_line += 1

wb.save("Round Tube4 Complete.xlsx")


# round tube2 for K128
# w1
# 