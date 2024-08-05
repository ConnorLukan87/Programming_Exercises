from openpyxl import Workbook
from openpyxl import load_workbook


# get all the descriptions from each worksheet, and hold a mapping to it's columns A-S

num_rows = [1662, 303, 45, 75, 45]

non_duplicate_descriptions = set()

cols = "ABCDEFGHIJKLMNOPQRS"
write_file = load_workbook('test_0.xlsx')
write_ws = write_file["RT=Round Tube"]
write_line = 4


for wb_i in range(5):
    name = "Round Tube" + str(wb_i) + " Complete.xlsx"
    wb = load_workbook(name)
    ws = wb.active
    # read all the descriptinos. if the current description is not in the set of all descriptions, add the entire row
    for line in range(4, num_rows[wb_i]):
        if ws["A" + str(line)].value not in non_duplicate_descriptions:
            non_duplicate_descriptions.add(ws["A" + str(line)].value)
            for col in cols:
                write_ws[col + str(write_line)] = ws[col + str(line)].value
            write_line += 1

write_file.save("Round Tube Complete.xlsx")