from openpyxl import Workbook
from openpyxl import load_workbook
import fractions
import math

def is_mixed(number):
    return '-' in line

def mixed_to_float(s):
    one, two = s.split('-')
    num, denom = two.split('/')
    num = int(num)
    denom = int(denom)
    num = num + denom*int(one)
    return num / denom

def is_int(number):
    return '/' not in number and '.' not in number

def is_float(number):
    return '/' not in number

data_file = open('flat_bar2.txt', 'r')

wb = load_workbook("test_0.xlsx")
ws = wb["FB=Flat Bar"]
#["FB=Flat Bar"]

current_thickness = 0
current_thickness_str = ""
desc_set = set()
current_line = 4
for line in data_file.readlines():
    numbers = line.strip().split()
    if len(numbers) == 0:
        continue
    elif len(numbers) == 1: # new thickness
        if is_mixed(numbers[0]):
            current_thickness = mixed_to_float(numbers[0])
            one, two = numbers[0].split('-')
            current_thickness_str = one + " " + two
        elif is_int(numbers[0]) or is_float(numbers[0]):
            current_thickness = float(numbers[0])
            current_thickness_str = numbers[0]
        else:
            num, denom = numbers[0].split('/')
            current_thickness = int(num)/int(denom)
            current_thickness_str = numbers[0]
        continue

    # read the line
    desc = "Flat Bar " + current_thickness_str + " x "

    if is_mixed(numbers[0]):
        one, two = numbers[0].split('-')
        desc += one + " " + two
    else:
        desc += numbers[0]

    if desc not in desc_set:
        desc_set.add(desc)
        ws["A" + str(current_line)] = desc
        ws["D" + str(current_line)] = current_thickness
        w1 = 0
        if is_mixed(numbers[0]):
            w1 = mixed_to_float(numbers[0])
        elif '/' in numbers[0]:
            n, d = numbers[0].split('/')
            w1 = float(n)/float(d)
        else:
            w1 = float(numbers[0])

        ws["F" + str(current_line)] = w1
        ws["H" + str(current_line)] = numbers[1]
        current_line += 1

wb.save("Flat Bar Complete.xlsx")

