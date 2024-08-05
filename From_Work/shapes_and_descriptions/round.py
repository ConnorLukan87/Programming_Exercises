from openpyxl import Workbook
from openpyxl import load_workbook
from simplify import simplify
from fractions import Fraction


wb = load_workbook('round.xlsx')

sheetnames = wb.sheetnames

# assume we have a 3d array of data sheets x descriptions x columns
data_arr = {name : [[]] for name in sheetnames}

# read in the flat bar formatted text file
data_file = open('round.txt', 'r')

# w1 w2 th1 th2 wgt size_no
ws = wb["R=Round"]

# 3/16 .1036 1.243

current_line = 4
data_arr = [[]]
whole_part = 0
desc_set = set()
for line in data_file.readlines():
    numbers = line.strip().split()

    if len(numbers) == 0:
        continue
    elif len(numbers) == 1:
        whole_part = numbers[0]
        continue
    
    elif len(numbers) == 2: # new whole number part
        weight_flag = numbers[1]
        whole_part = numbers[0]
        continue
    
    if "/" not in numbers[0]:
        frac = Fraction(float(numbers[0]))
        num = frac.numerator
        denom = frac.denominator
    else:
        num, denom = numbers[0].split('/')

    num = int(num)
    denom = int(denom)

    mixed_num = simplify(num + denom *int(whole_part), denom)

    desc = "Round " + mixed_num
   
    if desc not in desc_set:
        ws["I" + str(current_line)] = weight_flag
        ws["F" + str(current_line)] = int(whole_part) + (num / denom)
        ws["H" + str(current_line)] = numbers[1]
        ws["A"+ str(current_line)] = desc
        current_line+=1
        desc_set.add(desc)
    


wb.save("Round Complete.xlsx")
