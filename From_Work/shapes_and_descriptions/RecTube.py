from openpyxl import Workbook
from openpyxl import load_workbook
import fractions
import math

def is_mixed(number):
    return '-' in number

def mixed_to_float(s):
    one, two = s.split('-')
    num, denom = two.split('/')
    num = int(num)
    denom = int(denom)
    num = num + denom*int(one)
    return num / denom

def is_int(number):
    return '/' not in number and '.' not in number

def is_float(number):
    return '/' not in number

data_file = open('rectangular_tube.txt', 'r')

wb = load_workbook("test_0.xlsx")
ws = wb["RCT=Rectangular Tube"]
#["FB=Flat Bar"]

current_w1 = 0
current_w2 = 0
current_w1_str = ""
current_w2_str = ""
desc_set = set()
current_line = 4
for line in data_file.readlines():
    numbers = line.strip().split()
    if len(numbers) == 0:
        continue
    elif 'x' in numbers: # new w1 x w2
        if is_mixed(numbers[0]):
            current_w1 = mixed_to_float(numbers[0])
            one, two = numbers[0].split('-')
            current_w1_str = one + " " + two
        elif is_int(numbers[0]) or is_float(numbers[0]):
            current_w1 = float(numbers[0])
            current_w1_str = numbers[0]
        else:
            num, denom = numbers[0].split('/')
            current_w1 = int(num)/int(denom)
            current_w1_str = numbers[0]
        
        if is_mixed(numbers[2]):
            current_w2 = mixed_to_float(numbers[2])
            one, two = numbers[2].split('-')
            current_w2_str = one + " " + two
        elif is_int(numbers[2]) or is_float(numbers[2]):
            current_w2 = float(numbers[2])
            current_w2_str = numbers[2]
        else:
            num, denom = numbers[2].split('/')
            current_w2 = int(num)/int(denom)
            current_w2_str = numbers[2]

        continue

    # read the line (thickness, wgt)
    desc = "Rectangular Tube " + current_w1_str + " x " + current_w2_str + " x "

    if is_mixed(numbers[0]):
        print(numbers[0])
        one, two = numbers[0].split('-')
        desc += one + " " + two
    else:
        desc += numbers[0]

    if desc not in desc_set:
        desc_set.add(desc)
        ws["A" + str(current_line)] = desc
        ws["D" + str(current_line)] = float(numbers[0])
        ws["F" + str(current_line)] = current_w1
        ws["G" + str(current_line)] = current_w2
        ws["H" + str(current_line)] = numbers[1]
        current_line += 1

wb.save("Rectangular Tube Complete.xlsx")

