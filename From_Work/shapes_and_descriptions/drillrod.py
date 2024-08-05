from openpyxl import Workbook
from openpyxl import load_workbook
from simplify import simplify
from fractions import Fraction


wb = load_workbook('drillrod.xlsx')

sheetnames = wb.sheetnames

# assume we have a 3d array of data sheets x descriptions x columns
data_arr = {name : [[]] for name in sheetnames}

# read in the flat bar formatted text file
data_file = open('drillrod.txt', 'r')

# w1 w2 th1 th2 wgt size_no
ws = wb["DR=Drill Rod"]

# 3/16 .1036 1.243

current_line = 4
data_arr = [[]]
whole_part = 0
desc_set = set()
for line in data_file.readlines():
    numbers = line.strip().split()

    if len(numbers) == 0:
        continue

    desc = "Square " + numbers[1]
   
    if desc not in desc_set:
        ws["A"+ str(current_line)] = desc                               # description
        ws["B" + str(current_line)] = "S"                               # category
        ws["C" + str(current_line)] = "DR"                              # shape type
        ws["F" + str(current_line)] = numbers[0]                        # width 1
        ws["H" + str(current_line)] = numbers[2]                        # wgt.
        ws["I" + str(current_line)] = "3L"                              # wgt. type   
        ws["J" + str(current_line)] =  ws["J4"].value                   # color marks
        ws["K" + str(current_line)] = "Y"                               # restrict size entry
        
        current_line+=1
        desc_set.add(desc)


wb.save("Drill Rod Complete.xlsx")
